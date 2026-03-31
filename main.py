from __future__ import annotations

import argparse
import json
import locale as pylocale
import os
import sys
import time
import tracemalloc
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Callable, Optional

import tkinter as tk
import tkinter.font as tkfont
from tkinter import colorchooser, filedialog, messagebox, ttk


def get_app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


@dataclass
class AppConfig:
    data_file: str = "scores_data.json"
    theme: str = "light"
    font_scale: float = 1.0
    history_render_limit: int = 200
    save_debounce_ms: int = 250
    responsive_breakpoint_px: int = 980
    locale: Optional[str] = None
    enable_emoji: bool = True


@dataclass(frozen=True)
class Theme:
    bg_primary: str
    bg_secondary: str
    bg_surface: str
    text_primary: str
    text_secondary: str
    border: str
    accent_blue: str
    accent_green: str
    accent_red: str
    accent_purple: str
    accent_neutral: str


LIGHT_THEME = Theme(
    bg_primary="#f5f7fa",
    bg_secondary="#ffffff",
    bg_surface="#f8f9fa",
    text_primary="#2c3e50",
    text_secondary="#7f8c8d",
    border="#dcdfe6",
    accent_blue="#3498db",
    accent_green="#2ecc71",
    accent_red="#e74c3c",
    accent_purple="#9b59b6",
    accent_neutral="#34495e",
)

DARK_THEME = Theme(
    bg_primary="#0f141a",
    bg_secondary="#141b24",
    bg_surface="#1b2430",
    text_primary="#e6eef8",
    text_secondary="#9fb0c0",
    border="#2a3646",
    accent_blue="#4aa3ff",
    accent_green="#45d483",
    accent_red="#ff6b5e",
    accent_purple="#c08cff",
    accent_neutral="#6b7c8e",
)

GROUP_COLOR_PALETTE: list[dict[str, str]] = [
    {"hex": "#1f77b4", "name": "蓝色", "badge": "A", "pattern": "diag"},
    {"hex": "#ff7f0e", "name": "橙色", "badge": "B", "pattern": "cross"},
    {"hex": "#2ca02c", "name": "绿色", "badge": "C", "pattern": "dots"},
    {"hex": "#d62728", "name": "红色", "badge": "D", "pattern": "hatch"},
    {"hex": "#9467bd", "name": "紫色", "badge": "E", "pattern": "diag"},
    {"hex": "#17becf", "name": "青色", "badge": "F", "pattern": "cross"},
    {"hex": "#8c564b", "name": "棕色", "badge": "G", "pattern": "dots"},
    {"hex": "#e377c2", "name": "粉色", "badge": "H", "pattern": "hatch"},
    {"hex": "#005f73", "name": "深青", "badge": "I", "pattern": "diag"},
    {"hex": "#9a031e", "name": "深红", "badge": "J", "pattern": "cross"},
    {"hex": "#264653", "name": "深蓝", "badge": "K", "pattern": "dots"},
    {"hex": "#6d597a", "name": "灰紫", "badge": "L", "pattern": "hatch"},
]

COLOR_CHOICES = [f"{p['badge']} {p['name']} {p['hex']}" for p in GROUP_COLOR_PALETTE]


def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    s = hex_color.strip()
    if not s.startswith("#"):
        raise ValueError("invalid color")
    s = s[1:]
    if len(s) != 6:
        raise ValueError("invalid color")
    return int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)


def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    r, g, b = rgb
    return f"#{max(0, min(255, r)):02x}{max(0, min(255, g)):02x}{max(0, min(255, b)):02x}"


def _blend_hex(a: str, b: str, t: float) -> str:
    ar, ag, ab = _hex_to_rgb(a)
    br, bg, bb = _hex_to_rgb(b)
    rr = int(round(ar + (br - ar) * t))
    rg = int(round(ag + (bg - ag) * t))
    rb = int(round(ab + (bb - ab) * t))
    return _rgb_to_hex((rr, rg, rb))


def _is_light(hex_color: str) -> bool:
    r, g, b = _hex_to_rgb(hex_color)
    return (0.2126 * r + 0.7152 * g + 0.0722 * b) >= 165


def _srgb_channel_to_linear(c: float) -> float:
    if c <= 0.04045:
        return c / 12.92
    return ((c + 0.055) / 1.055) ** 2.4


def _relative_luminance(hex_color: str) -> float:
    r, g, b = _hex_to_rgb(hex_color)
    rs = _srgb_channel_to_linear(r / 255.0)
    gs = _srgb_channel_to_linear(g / 255.0)
    bs = _srgb_channel_to_linear(b / 255.0)
    return 0.2126 * rs + 0.7152 * gs + 0.0722 * bs


def _contrast_ratio(fg_hex: str, bg_hex: str) -> float:
    l1 = _relative_luminance(fg_hex)
    l2 = _relative_luminance(bg_hex)
    lighter = max(l1, l2)
    darker = min(l1, l2)
    return (lighter + 0.05) / (darker + 0.05)


def _best_text_color(bg_hex: str) -> str:
    dark = "#111111"
    light = "#ffffff"
    return dark if _contrast_ratio(dark, bg_hex) >= _contrast_ratio(light, bg_hex) else light


def _extract_hex(s: str) -> str | None:
    for token in reversed(str(s).replace("，", " ").replace(",", " ").split()):
        if token.startswith("#") and len(token) == 7:
            try:
                _hex_to_rgb(token)
                return token.lower()
            except Exception:
                continue
    if str(s).startswith("#") and len(str(s).strip()) == 7:
        try:
            _hex_to_rgb(str(s).strip())
            return str(s).strip().lower()
        except Exception:
            return None
    return None


def _next_palette_entry(used_hex: set[str]) -> dict[str, str]:
    for p in GROUP_COLOR_PALETTE:
        if p["hex"].lower() not in used_hex:
            return p
    idx = len(used_hex) % len(GROUP_COLOR_PALETTE)
    return GROUP_COLOR_PALETTE[idx]


class Tooltip:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self._tip: tk.Toplevel | None = None
        self._after_id: str | None = None

    def show(self, widget: tk.Widget, text: str, *, delay_ms: int = 250) -> None:
        self.hide()

        def _do() -> None:
            if self._tip is not None:
                return
            tip = tk.Toplevel(self.root)
            tip.wm_overrideredirect(True)
            tip.attributes("-topmost", True)
            label = tk.Label(
                tip,
                text=text,
                bg="#111111",
                fg="#ffffff",
                font=("Segoe UI", 10),
                padx=10,
                pady=6,
                relief=tk.SOLID,
                bd=1,
            )
            label.pack()
            x = widget.winfo_rootx() + 10
            y = widget.winfo_rooty() + widget.winfo_height() + 8
            tip.wm_geometry(f"+{x}+{y}")
            self._tip = tip

        self._after_id = self.root.after(delay_ms, _do)

    def hide(self) -> None:
        if self._after_id is not None:
            try:
                self.root.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None
        if self._tip is not None:
            try:
                self._tip.destroy()
            except Exception:
                pass
            self._tip = None

STRINGS: dict[str, dict[str, str]] = {
    "zh_CN": {
        "app_title": "小组积分管理系统",
        "menu_group": "小组管理",
        "menu_help": "帮助",
        "menu_manage_groups": "管理小组",
        "menu_add_group": "添加小组",
        "menu_usage": "使用说明",
        "rank_title": "🏆 实时排名",
        "history_title": "📋 历史记录",
        "batch_title": "⚙️ 批量操作",
        "btn_clear": "清空",
        "btn_apply": "应用",
        "btn_reset_scores": "重置所有积分",
        "batch_group": "小组：",
        "batch_score": "分数：",
        "all_groups": "所有小组",
        "err_invalid_number": "请输入有效的数字！",
        "confirm": "确认",
        "info": "提示",
        "success": "成功",
        "err": "错误",
        "confirm_reset": "确定要重置所有小组积分吗？此操作不可恢复！",
        "confirm_clear_history": "确定要清空所有历史记录吗？此操作不可恢复！",
        "help_title": "使用说明",
        "help_text": """小组积分管理系统 - 使用说明

【基本操作】
• 点击小组卡片上的快捷按钮快速加减分
• 使用自定义输入框设置任意分数
• 查看右侧实时排名和历史记录

【小组管理】
• 菜单 → 小组管理 → 管理小组
• 可以添加、编辑、删除小组
• 可以为每个小组设置颜色和名称

【成员管理】
• 菜单 → 小组管理 → 管理小组
• 选择小组后点击\"管理成员\"
• 可以添加、编辑、删除小组成员
• 支持设置角色：组长、副组长、组员

【批量操作】
• 在右侧选择单个或所有小组
• 使用快捷按钮或自定义分数批量操作

【数据安全】
• 所有操作自动保存到 scores_data.json
• 删除操作会有确认提示""",
    },
    "en_US": {
        "app_title": "Group Score Manager",
        "menu_group": "Groups",
        "menu_help": "Help",
        "menu_manage_groups": "Manage Groups",
        "menu_add_group": "Add Group",
        "menu_usage": "Usage",
        "rank_title": "🏆 Ranking",
        "history_title": "📋 History",
        "batch_title": "⚙️ Batch",
        "btn_clear": "Clear",
        "btn_apply": "Apply",
        "btn_reset_scores": "Reset All Scores",
        "batch_group": "Group:",
        "batch_score": "Score:",
        "all_groups": "All Groups",
        "err_invalid_number": "Please enter a valid number.",
        "confirm": "Confirm",
        "info": "Info",
        "success": "Success",
        "err": "Error",
        "confirm_reset": "Reset all group scores? This cannot be undone.",
        "confirm_clear_history": "Clear all history? This cannot be undone.",
        "help_title": "Usage",
        "help_text": "Group Score Manager\n\nUse buttons to change scores, view ranking and history.",
    },
}


def resolve_locale(explicit: Optional[str]) -> str:
    return "zh_CN"


def resolve_theme(name: str) -> Theme:
    return DARK_THEME if name.lower() == "dark" else LIGHT_THEME


class GroupScoreApp:
    def __init__(self, root: tk.Tk, config: Optional[AppConfig] = None):
        self.root = root
        self.config = config or AppConfig()
        self.locale = resolve_locale(self.config.locale)
        self.strings = STRINGS.get(self.locale, STRINGS["zh_CN"])
        self.theme = resolve_theme(self.config.theme)

        self._save_after_id: Optional[str] = None
        self._layout_after_id: Optional[str] = None
        self._history_dirty = True
        self._ranking_dirty = True
        self._scores_dirty = True
        self._cards_ready = False
        self._right_ready = False
        self._layout_mode: str = "side"
        self._cards_columns: int = 2

        self.rank_listbox: Optional[tk.Listbox] = None
        self.history_text: Optional[tk.Text] = None
        self.batch_group_combo: Optional[ttk.Combobox] = None
        self.batch_group_var: Optional[tk.StringVar] = None
        self.batch_score_entry: Optional[tk.Entry] = None

        self._configure_scaling()
        self._configure_fonts()
        self._configure_styles()

        self.root.title(self.t("app_title"))
        self._configure_window_defaults()
        self.root.configure(bg=self.theme.bg_primary)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        self.data_file = self.config.data_file
        self.load_data()
        self.tooltip = Tooltip(self.root)

        self._build_shell()
        self._disable_main_window_scroll()
        self.create_menu()

        self.root.after(1, self._init_right_panel_progressively)
        self.root.after(1, self._build_group_cards_progressively)
        self.root.after(1, self.refresh_all)

    def _disable_main_window_scroll(self) -> None:
        def handler(event: tk.Event) -> str | None:
            try:
                top = event.widget.winfo_toplevel()
            except Exception:
                return None
            return "break" if top == self.root else None

        for seq in ("<MouseWheel>", "<Shift-MouseWheel>", "<Button-4>", "<Button-5>"):
            try:
                self.root.bind_all(seq, handler, add="+")
            except Exception:
                pass

    def t(self, key: str) -> str:
        return self.strings.get(key, key)

    def _configure_scaling(self) -> None:
        try:
            ppi = float(self.root.winfo_fpixels("1i"))
            self.root.tk.call("tk", "scaling", ppi / 72.0)
        except Exception:
            pass

    def _configure_fonts(self) -> None:
        family = "微软雅黑" if self.locale == "zh_CN" else "Segoe UI"
        s = max(0.5, float(self.config.font_scale))
        self.font_body = tkfont.Font(family=family, size=int(round(10 * s)))
        self.font_small = tkfont.Font(family=family, size=int(round(9 * s)))
        self.font_title = tkfont.Font(family=family, size=int(round(11 * s)), weight="bold")
        self.font_header = tkfont.Font(family=family, size=int(round(12 * s)), weight="bold")
        self.font_score = tkfont.Font(family="Arial Black", size=int(round(34 * s)), weight="bold")

    def _configure_styles(self) -> None:
        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except Exception:
            pass
        self.style.configure("TLabelframe", background=self.theme.bg_primary)
        self.style.configure(
            "TLabelframe.Label",
            background=self.theme.bg_primary,
            foreground=self.theme.text_primary,
            font=self.font_title,
        )

    def _configure_window_defaults(self) -> None:
        try:
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            w = max(900, int(sw * 0.75))
            h = max(560, int(sh * 0.75))
            self.root.geometry(f"{w}x{h}")
            self.root.minsize(900, 560)
        except Exception:
            self.root.geometry("1200x700")

    def _build_shell(self) -> None:
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.main_container = tk.Frame(self.root, bg=self.theme.bg_primary)
        self.main_container.grid(row=0, column=0, sticky="nsew")
        self.main_container.rowconfigure(0, weight=1)
        self.main_container.columnconfigure(0, weight=1)
        self.main_container.columnconfigure(1, weight=0)

        self.left_container = tk.Frame(self.main_container, bg=self.theme.bg_primary)
        self.right_container = tk.Frame(self.main_container, bg=self.theme.bg_primary)
        self.left_container.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.right_container.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        self.cards_placeholder = tk.Label(
            self.left_container,
            text="…",
            bg=self.theme.bg_primary,
            fg=self.theme.text_secondary,
            font=self.font_body,
        )
        self.cards_placeholder.pack(fill=tk.BOTH, expand=True)

        self.right_placeholder = tk.Label(
            self.right_container,
            text="…",
            bg=self.theme.bg_primary,
            fg=self.theme.text_secondary,
            font=self.font_body,
        )
        self.right_placeholder.pack(fill=tk.BOTH, expand=True)

        self.root.bind("<Configure>", self._on_root_configure)
        self._apply_responsive_layout(self.root.winfo_width())

    def _on_root_configure(self, event: tk.Event) -> None:
        if self._layout_after_id is not None:
            self.root.after_cancel(self._layout_after_id)
        self._layout_after_id = self.root.after(80, lambda: self._apply_responsive_layout(self.root.winfo_width()))

    def _apply_responsive_layout(self, width: int) -> None:
        mode = "stack" if width < int(self.config.responsive_breakpoint_px) else "side"
        if mode != self._layout_mode:
            self._layout_mode = mode
            if mode == "side":
                self.main_container.columnconfigure(1, weight=0, minsize=320)
                self.right_container.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
                self.left_container.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
            else:
                self.main_container.columnconfigure(1, weight=0, minsize=0)
                self.right_container.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
                self.left_container.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
                self.main_container.rowconfigure(1, weight=0)
        self._update_cards_columns()

    @property
    def layout_mode(self) -> str:
        return self._layout_mode

    def load_data(self) -> None:
        if os.path.exists(self.data_file):
            with open(self.data_file, "r", encoding="utf-8") as f:
                self.data = json.load(f)
        else:
            self.data = {
                "groups": [
                    {"name": "第一小组", "score": 0.0, "color": GROUP_COLOR_PALETTE[0]["hex"], "members": []},
                    {"name": "第二小组", "score": 0.0, "color": GROUP_COLOR_PALETTE[1]["hex"], "members": []},
                    {"name": "第三小组", "score": 0.0, "color": GROUP_COLOR_PALETTE[2]["hex"], "members": []},
                    {"name": "第四小组", "score": 0.0, "color": GROUP_COLOR_PALETTE[3]["hex"], "members": []},
                    {"name": "第五小组", "score": 0.0, "color": GROUP_COLOR_PALETTE[4]["hex"], "members": []},
                    {"name": "第六小组", "score": 0.0, "color": GROUP_COLOR_PALETTE[5]["hex"], "members": []},
                ],
                "history": [],
            }
            self.save_data()

        if "groups" not in self.data or not isinstance(self.data["groups"], list):
            self.data["groups"] = []
        if "history" not in self.data or not isinstance(self.data["history"], list):
            self.data["history"] = []
        if len(self.data.get("groups", [])) < 6:
            cn = ["一", "二", "三", "四", "五", "六"]
            added = False
            while len(self.data["groups"]) < 6:
                idx = len(self.data["groups"])
                self.data["groups"].append(
                    {"name": f"第{cn[idx]}小组", "score": 0.0, "color": GROUP_COLOR_PALETTE[idx]["hex"], "members": []}
                )
                added = True
            if added:
                self.save_data()
        used_hex: set[str] = set()
        used_badges: set[str] = set()
        for group in self.data.get("groups", []):
            if "members" not in group or not isinstance(group.get("members"), list):
                group["members"] = []
            color_raw = group.get("color", "")
            color_hex = _extract_hex(color_raw) or GROUP_COLOR_PALETTE[0]["hex"]
            if color_hex.lower() in used_hex:
                color_hex = _next_palette_entry(used_hex)["hex"]
            group["color"] = color_hex
            used_hex.add(color_hex.lower())

        for group in self.data.get("groups", []):
            if "badge" not in group or not str(group.get("badge", "")).strip():
                hex_color = str(group.get("color", GROUP_COLOR_PALETTE[0]["hex"])).lower()
                entry = next((p for p in GROUP_COLOR_PALETTE if p["hex"].lower() == hex_color), None) or GROUP_COLOR_PALETTE[0]
                badge = entry["badge"]
                if badge in used_badges:
                    badge = f"{badge}{len(used_badges) + 1}"
                group["badge"] = badge
            used_badges.add(str(group.get("badge")))
            if "pattern" not in group or not str(group.get("pattern", "")).strip():
                hex_color = str(group.get("color", GROUP_COLOR_PALETTE[0]["hex"])).lower()
                entry = next((p for p in GROUP_COLOR_PALETTE if p["hex"].lower() == hex_color), None)
                group["pattern"] = (entry or GROUP_COLOR_PALETTE[0])["pattern"]

    def save_data(self) -> None:
        with open(self.data_file, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def _mark_dirty(self) -> None:
        if self._save_after_id is not None:
            return
        self._save_after_id = self.root.after(self.config.save_debounce_ms, self.flush_save)

    def flush_save(self) -> None:
        if self._save_after_id is not None:
            try:
                self.root.after_cancel(self._save_after_id)
            except Exception:
                pass
            self._save_after_id = None
        self.save_data()

    def on_close(self) -> None:
        try:
            self.flush_save()
        finally:
            self.root.destroy()

    def _init_right_panel_progressively(self) -> None:
        if self._right_ready:
            return
        if getattr(self, "right_placeholder", None) is not None:
            self.right_placeholder.destroy()
            self.right_placeholder = None
        self.init_ranking()
        self.root.after(1, self._init_right_panel_step_2)

    def _init_right_panel_step_2(self) -> None:
        self.init_history()
        self.root.after(1, self._init_right_panel_step_3)

    def _init_right_panel_step_3(self) -> None:
        self.init_batch_ops()
        self._right_ready = True
        self._ranking_dirty = True
        self._history_dirty = True
        self.refresh_all()

    def init_ranking(self) -> None:
        rank_frame = tk.LabelFrame(self.right_container, text=self.t("rank_title"), font=self.font_title,
                                   bg=self.theme.bg_primary, fg=self.theme.text_primary)
        rank_frame.pack(fill=tk.BOTH, padx=0, pady=(0, 10), expand=False)

        self.rank_listbox = tk.Listbox(
            rank_frame,
            height=8,
            bg=self.theme.bg_secondary,
            fg=self.theme.text_primary,
            highlightthickness=1,
            highlightbackground=self.theme.border,
            font=self.font_body,
        )
        self.rank_listbox.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

    def init_history(self) -> None:
        hist_frame = tk.LabelFrame(self.right_container, text=self.t("history_title"), font=self.font_title,
                                   bg=self.theme.bg_primary, fg=self.theme.text_primary)
        hist_frame.pack(fill=tk.BOTH, padx=0, pady=(0, 10), expand=True)

        clear_btn = tk.Button(hist_frame, text=self.t("btn_clear"), command=self.clear_history,
                              bg=self.theme.bg_surface, fg=self.theme.text_primary, bd=0, font=self.font_small)
        clear_btn.pack(anchor="e", padx=6, pady=(6, 0))

        self.history_text = tk.Text(
            hist_frame,
            height=10,
            width=35,
            bg=self.theme.bg_secondary,
            fg=self.theme.text_primary,
            font=self.font_small,
            bd=1,
            relief=tk.SOLID,
            highlightthickness=0,
            wrap=tk.WORD,
        )
        self.history_text.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.history_text.config(state=tk.DISABLED)
        self._configure_history_tags()

    def _configure_history_tags(self) -> None:
        if not self.history_text:
            return
        self.history_text.tag_config("time", foreground=self.theme.text_secondary, font=self.font_small)
        self.history_text.tag_config("pos", foreground=self.theme.accent_green, font=self.font_body)
        self.history_text.tag_config("neg", foreground=self.theme.accent_red, font=self.font_body)
        self.history_text.tag_config("zero", foreground=self.theme.text_secondary, font=self.font_body)
        self.history_text.tag_config("text", foreground=self.theme.text_primary, font=self.font_body)

    def init_batch_ops(self) -> None:
        batch_frame = tk.LabelFrame(self.right_container, text=self.t("batch_title"), font=self.font_title,
                                    bg=self.theme.bg_primary, fg=self.theme.text_primary)
        batch_frame.pack(fill=tk.BOTH, padx=0, pady=0, expand=False)

        tk.Label(batch_frame, text=self.t("batch_group"), bg=self.theme.bg_primary, fg=self.theme.text_primary,
                 font=self.font_body).grid(row=0, column=0, padx=6, pady=6, sticky="w")
        self.batch_group_var = tk.StringVar(value=self.t("all_groups"))
        self.batch_group_combo = ttk.Combobox(
            batch_frame,
            textvariable=self.batch_group_var,
            values=[self.t("all_groups")] + [g.get("name", "") for g in self.data.get("groups", [])],
            state="readonly",
        )
        self.batch_group_combo.grid(row=0, column=1, padx=6, pady=6, sticky="ew")

        tk.Label(batch_frame, text=self.t("batch_score"), bg=self.theme.bg_primary, fg=self.theme.text_primary,
                 font=self.font_body).grid(row=1, column=0, padx=6, pady=6, sticky="w")
        self.batch_score_entry = tk.Entry(batch_frame, width=10, font=self.font_body, bg=self.theme.bg_secondary,
                                          fg=self.theme.text_primary, bd=1, relief=tk.SOLID,
                                          highlightbackground=self.theme.border, highlightthickness=1)
        self.batch_score_entry.grid(row=1, column=1, padx=6, pady=6, sticky="ew")

        btn_frame = tk.Frame(batch_frame, bg=self.theme.bg_primary)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=(6, 10))
        btn_style = {"font": self.font_body, "bd": 0, "padx": 10, "pady": 6, "bg": self.theme.bg_surface,
                     "fg": self.theme.text_primary}
        tk.Button(btn_frame, text="+1", command=lambda: self.batch_change(1.0), **btn_style).pack(side="left", padx=3)
        tk.Button(btn_frame, text="-1", command=lambda: self.batch_change(-1.0), **btn_style).pack(side="left", padx=3)
        tk.Button(btn_frame, text="+0.5", command=lambda: self.batch_change(0.5), **btn_style).pack(side="left", padx=3)
        tk.Button(btn_frame, text="-0.5", command=lambda: self.batch_change(-0.5), **btn_style).pack(side="left", padx=3)

        tk.Button(batch_frame, text=self.t("btn_apply"), command=self.batch_apply_custom,
                  bg=self.theme.accent_blue, fg="white", font=self.font_body, bd=0, pady=8).grid(
            row=3, column=0, columnspan=2, padx=6, pady=(0, 10), sticky="ew"
        )

        tk.Button(batch_frame, text=self.t("btn_reset_scores"), bg=self.theme.accent_red, fg="white",
                  command=self.reset_all, font=self.font_body, bd=0, pady=8).grid(
            row=4, column=0, columnspan=2, padx=6, pady=(0, 6), sticky="ew"
        )

        batch_frame.columnconfigure(1, weight=1)

    def batch_change(self, delta: float) -> None:
        if not self.batch_group_var:
            return
        target = self.batch_group_var.get()
        for i, group in enumerate(self.data.get("groups", [])):
            if target == self.t("all_groups") or group.get("name") == target:
                self.data["groups"][i]["score"] = float(group.get("score", 0.0)) + float(delta)
                self.add_history(group.get("name", ""), delta)
        self._scores_dirty = True
        self._ranking_dirty = True
        self._mark_dirty()
        self.refresh_all()

    def batch_apply_custom(self) -> None:
        if not self.batch_group_var or not self.batch_score_entry:
            return
        try:
            val = float(self.batch_score_entry.get())
        except ValueError:
            messagebox.showerror(self.t("err"), self.t("err_invalid_number"))
            return
        target = self.batch_group_var.get()
        for i, group in enumerate(self.data.get("groups", [])):
            if target == self.t("all_groups") or group.get("name") == target:
                old = float(group.get("score", 0.0))
                self.data["groups"][i]["score"] = float(val)
                self.add_history(group.get("name", ""), float(val) - old)
        self.batch_score_entry.delete(0, tk.END)
        self._scores_dirty = True
        self._ranking_dirty = True
        self._mark_dirty()
        self.refresh_all()

    def reset_all(self) -> None:
        if messagebox.askyesno(self.t("confirm"), self.t("confirm_reset")):
            for i in range(len(self.data.get("groups", []))):
                self.data["groups"][i]["score"] = 0.0
            self.add_history("系统" if self.locale == "zh_CN" else "System",
                             "重置所有积分" if self.locale == "zh_CN" else "Reset all scores")
            self._scores_dirty = True
            self._ranking_dirty = True
            self._mark_dirty()
            self.refresh_all()

    def clear_history(self) -> None:
        if messagebox.askyesno(self.t("confirm"), self.t("confirm_clear_history")):
            self.data["history"] = []
            self._history_dirty = True
            self._mark_dirty()
            self.refresh_history()

    def _build_group_cards_progressively(self) -> None:
        if getattr(self, "cards_frame", None) is not None:
            return
        if getattr(self, "cards_placeholder", None) is not None:
            self.cards_placeholder.destroy()
            self.cards_placeholder = None

        self.cards_frame = tk.Frame(self.left_container, bg=self.theme.bg_primary)
        self.cards_frame.pack(fill=tk.BOTH, expand=True)

        self.group_cards: list[dict[str, Any]] = []
        self._create_cards_chunk(0)

    def _create_cards_chunk(self, start_index: int) -> None:
        chunk_size = 2
        groups = self.data.get("groups", [])
        end = min(len(groups), start_index + chunk_size)
        for i in range(start_index, end):
            self._create_group_card(i)
        self._layout_cards()
        if end < len(groups):
            self.root.after(1, lambda: self._create_cards_chunk(end))
            return
        self._cards_ready = True
        self._scores_dirty = True
        self.refresh_group_cards()

    def _draw_swatch(self, canvas: tk.Canvas, *, color: str, pattern: str, badge: str) -> None:
        canvas.delete("all")
        canvas.configure(bg=color)
        fg = _best_text_color(color)
        w = int(canvas.cget("width"))
        h = int(canvas.cget("height"))
        if pattern == "diag":
            step = 6
            for x in range(-h, w + h, step):
                canvas.create_line(x, 0, x + h, h, fill=fg, width=1, stipple="gray50")
        elif pattern == "cross":
            step = 7
            for x in range(0, w + step, step):
                canvas.create_line(x, 0, x, h, fill=fg, width=1, stipple="gray50")
            for y in range(0, h + step, step):
                canvas.create_line(0, y, w, y, fill=fg, width=1, stipple="gray50")
        elif pattern == "dots":
            step = 7
            r = 1
            for x in range(3, w, step):
                for y in range(3, h, step):
                    canvas.create_oval(x - r, y - r, x + r, y + r, fill=fg, outline=fg)
        else:
            step = 7
            for x in range(-h, w + h, step):
                canvas.create_line(x, 0, x + h, h, fill=fg, width=1)
                canvas.create_line(x, h, x + h, 0, fill=fg, width=1)
        canvas.create_text(w // 2, h // 2, text=badge, fill=fg, font=self.font_title)

    def _hover_text_for_group(self, group: dict[str, Any]) -> str:
        name = str(group.get("name", ""))
        badge = str(group.get("badge", ""))
        color = str(group.get("color", ""))
        return f"小组：{name}\n标识：{badge}\n颜色：{color}"

    def _bind_tooltip(self, widget: tk.Widget, text_getter: Callable[[], str]) -> None:
        widget.bind("<Enter>", lambda _e: self.tooltip.show(widget, text_getter()))
        widget.bind("<Leave>", lambda _e: self.tooltip.hide())

    def _create_group_card(self, group_idx: int) -> None:
        group = self.data["groups"][group_idx]
        border_color = str(group.get("color", self.theme.accent_blue))
        badge = str(group.get("badge", str(group_idx + 1)))
        pattern = str(group.get("pattern", "diag"))
        title_fg = _best_text_color(border_color)

        card = tk.Frame(
            self.cards_frame,
            bg=self.theme.bg_secondary,
            highlightbackground=border_color,
            highlightcolor=border_color,
            highlightthickness=3,
            relief=tk.RAISED,
            bd=0,
        )

        title_frame = tk.Frame(card, bg=border_color)
        title_frame.pack(fill=tk.X)
        swatch = tk.Canvas(title_frame, width=34, height=22, highlightthickness=0, bd=0)
        swatch.pack(side=tk.LEFT, padx=10, pady=8)
        self._draw_swatch(swatch, color=border_color, pattern=pattern, badge=badge)
        title_label = tk.Label(title_frame, text=f"{badge}  {str(group.get('name', ''))}", bg=border_color, fg=title_fg,
                               font=self.font_header, pady=8)
        title_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self._bind_tooltip(title_frame, lambda g=group: self._hover_text_for_group(g))
        self._bind_tooltip(swatch, lambda g=group: self._hover_text_for_group(g))
        self._bind_tooltip(title_label, lambda g=group: self._hover_text_for_group(g))

        def on_enter(_e: tk.Event) -> None:
            try:
                card.configure(highlightthickness=4)
            except Exception:
                pass

        def on_leave(_e: tk.Event) -> None:
            try:
                card.configure(highlightthickness=3)
            except Exception:
                pass

        card.bind("<Enter>", on_enter)
        card.bind("<Leave>", on_leave)

        content_frame = tk.Frame(card, bg=self.theme.bg_secondary)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        score_var = tk.StringVar(value=f"{float(group.get('score', 0.0)):.1f}")
        score_label = tk.Label(content_frame, textvariable=score_var, bg=self.theme.bg_secondary, fg=border_color,
                               font=self.font_score)
        score_label.pack(pady=(6, 0))

        score_hint = tk.Label(content_frame, text="当前积分" if self.locale == "zh_CN" else "Score",
                              bg=self.theme.bg_secondary, fg=self.theme.text_secondary, font=self.font_small)
        score_hint.pack(pady=(0, 8))

        btn_frame = tk.Frame(content_frame, bg=self.theme.bg_secondary)
        btn_frame.pack(pady=8)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 10, "pady": 6}
        tk.Button(btn_frame, text="+1", bg=self.theme.accent_green, fg="white",
                  command=lambda g=group_idx: self.change_score(g, 1.0), **btn_style).grid(row=0, column=0, padx=4, pady=2)
        tk.Button(btn_frame, text="-1", bg=self.theme.accent_red, fg="white",
                  command=lambda g=group_idx: self.change_score(g, -1.0), **btn_style).grid(row=0, column=1, padx=4, pady=2)
        tk.Button(btn_frame, text="+0.5", bg=self.theme.accent_blue, fg="white",
                  command=lambda g=group_idx: self.change_score(g, 0.5), **btn_style).grid(row=0, column=2, padx=4, pady=2)
        tk.Button(btn_frame, text="-0.5", bg=self.theme.accent_purple, fg="white",
                  command=lambda g=group_idx: self.change_score(g, -0.5), **btn_style).grid(row=0, column=3, padx=4, pady=2)

        self.group_cards.append(
            {
                "group_idx": group_idx,
                "card": card,
                "score_var": score_var,
                "title_frame": title_frame,
                "title_label": title_label,
                "score_label": score_label,
                "swatch": swatch,
                "last_color": border_color,
                "anim_after_id": None,
                "animating": False,
            }
        )

    def _apply_group_color_to_card(self, card_info: dict[str, Any], *, color: str) -> None:
        fg = _best_text_color(color)
        idx = int(card_info["group_idx"])
        if idx >= len(self.data.get("groups", [])):
            return
        group = self.data["groups"][idx]
        badge = str(group.get("badge", ""))
        pattern = str(group.get("pattern", "diag"))
        name = str(group.get("name", ""))
        try:
            card_info["card"].configure(highlightbackground=color, highlightcolor=color)
        except Exception:
            pass
        try:
            card_info["title_frame"].configure(bg=color)
        except Exception:
            pass
        try:
            card_info["title_label"].configure(bg=color, fg=fg, text=f"{badge}  {name}")
        except Exception:
            pass
        try:
            card_info["score_label"].configure(fg=color)
        except Exception:
            pass
        try:
            self._draw_swatch(card_info["swatch"], color=color, pattern=pattern, badge=badge)
        except Exception:
            pass
        card_info["last_color"] = color

    def _animate_group_card_color(self, card_info: dict[str, Any], *, from_color: str, to_color: str) -> None:
        if str(from_color).lower() == str(to_color).lower():
            self._apply_group_color_to_card(card_info, color=to_color)
            return
        after_id = card_info.get("anim_after_id")
        if after_id is not None:
            try:
                self.root.after_cancel(after_id)
            except Exception:
                pass
        card_info["animating"] = True
        steps = 14
        duration_ms = 260
        step_ms = max(10, int(duration_ms / steps))

        def tick(i: int) -> None:
            t = min(1.0, max(0.0, i / steps))
            col = _blend_hex(from_color, to_color, t)
            self._apply_group_color_to_card(card_info, color=col)
            if i >= steps:
                card_info["anim_after_id"] = None
                card_info["animating"] = False
                self._apply_group_color_to_card(card_info, color=to_color)
                return
            card_info["anim_after_id"] = self.root.after(step_ms, lambda: tick(i + 1))

        tick(0)

    def _update_cards_columns(self) -> None:
        columns = 2
        if columns == self._cards_columns:
            return
        self._cards_columns = columns
        self._layout_cards()

    def _layout_cards(self) -> None:
        if not hasattr(self, "group_cards"):
            return
        for child in self.cards_frame.grid_slaves():
            child.grid_forget()
        cols = max(1, int(self._cards_columns))
        for i, card_info in enumerate(self.group_cards):
            row = i // cols
            col = i % cols
            card_info["card"].grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
        for c in range(cols):
            self.cards_frame.grid_columnconfigure(c, weight=1, uniform="cards")
        rows = max(1, (len(self.group_cards) + cols - 1) // cols)
        for r in range(rows):
            self.cards_frame.grid_rowconfigure(r, weight=1, uniform="cards")

    def change_score(self, group_idx: int, delta: float) -> None:
        self.data["groups"][group_idx]["score"] = float(self.data["groups"][group_idx].get("score", 0.0)) + float(delta)
        self.add_history(self.data["groups"][group_idx].get("name", ""), delta)
        self._scores_dirty = True
        self._ranking_dirty = True
        self._mark_dirty()
        self.refresh_all()

    def apply_custom_score(self, group_idx: int, entry: tk.Entry) -> None:
        try:
            val = float(entry.get())
        except ValueError:
            messagebox.showerror(self.t("err"), self.t("err_invalid_number"))
            return
        old_score = float(self.data["groups"][group_idx].get("score", 0.0))
        self.data["groups"][group_idx]["score"] = float(val)
        self.add_history(self.data["groups"][group_idx].get("name", ""), float(val) - old_score)
        entry.delete(0, tk.END)
        self._scores_dirty = True
        self._ranking_dirty = True
        self._mark_dirty()
        self.refresh_all()

    def add_history(self, group_name: str, change: Any) -> None:
        self.data["history"].append(
            {"group": group_name, "change": change, "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        )
        self._history_dirty = True
        self._mark_dirty()

    def create_menu(self) -> None:
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        group_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self.t("menu_group"), menu=group_menu)
        group_menu.add_command(label=self.t("menu_manage_groups"), command=self.open_group_manager)
        group_menu.add_command(
            label=("成员管理" if self.locale == "zh_CN" else "Members"),
            command=self.open_members_manager_global,
        )
        group_menu.add_separator()
        group_menu.add_command(label=self.t("menu_add_group"), command=lambda: self.add_new_group(None))

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=self.t("menu_help"), menu=help_menu)
        help_menu.add_command(label=self.t("menu_usage"), command=self.show_help)

    def show_help(self) -> None:
        messagebox.showinfo(self.t("help_title"), self.t("help_text"))

    def _rebuild_cards(self) -> None:
        if getattr(self, "cards_frame", None) is None:
            self._scores_dirty = True
            return
        for child in self.cards_frame.winfo_children():
            child.destroy()
        self.group_cards = []
        self._cards_ready = False
        self._create_cards_chunk(0)
        if self.batch_group_combo and self.batch_group_var:
            self.batch_group_combo.configure(values=[self.t("all_groups")] + [g.get("name", "") for g in self.data.get("groups", [])])
            if self.batch_group_var.get() not in self.batch_group_combo.cget("values"):
                self.batch_group_var.set(self.t("all_groups"))

    def open_group_manager(self) -> None:
        manager_win = tk.Toplevel(self.root)
        manager_win.title(self.t("menu_manage_groups"))
        manager_win.geometry("700x460")
        manager_win.configure(bg=self.theme.bg_primary)
        manager_win.transient(self.root)
        manager_win.grab_set()

        main_frame = tk.Frame(manager_win, bg=self.theme.bg_primary, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        list_frame = tk.LabelFrame(main_frame, text="小组列表" if self.locale == "zh_CN" else "Groups",
                                   font=self.font_title, bg=self.theme.bg_primary, fg=self.theme.text_primary)
        list_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("name", "members", "score")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=12)
        tree.heading("name", text="小组名称" if self.locale == "zh_CN" else "Name")
        tree.heading("members", text="成员数" if self.locale == "zh_CN" else "Members")
        tree.heading("score", text="当前积分" if self.locale == "zh_CN" else "Score")
        tree.column("name", width=260)
        tree.column("members", width=100, anchor=tk.CENTER)
        tree.column("score", width=120, anchor=tk.CENTER)

        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def reload_tree() -> None:
            for item in tree.get_children():
                tree.delete(item)
            for group in self.data.get("groups", []):
                member_count = len(group.get("members", []))
                tree.insert("", tk.END, values=(
                    group.get("name", ""),
                    f"{member_count}人" if self.locale == "zh_CN" else str(member_count),
                    f"{float(group.get('score', 0.0)):.1f}",
                ))

        def selected_group_index() -> Optional[int]:
            sel = tree.selection()
            if not sel:
                return None
            name = tree.item(sel[0]).get("values", [""])[0]
            for i, g in enumerate(self.data.get("groups", [])):
                if g.get("name") == name:
                    return i
            return None

        reload_tree()

        btn_frame = tk.Frame(main_frame, bg=self.theme.bg_primary)
        btn_frame.pack(fill=tk.X, pady=12)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 14, "pady": 8, "fg": "white"}

        def on_edit() -> None:
            idx = selected_group_index()
            if idx is None:
                messagebox.showwarning(self.t("info"), "请先选择要编辑的小组！" if self.locale == "zh_CN" else "Select a group.")
                return
            self.edit_group(idx, on_updated=reload_tree)

        def on_delete() -> None:
            idx = selected_group_index()
            if idx is None:
                messagebox.showwarning(self.t("info"), "请先选择要删除的小组！" if self.locale == "zh_CN" else "Select a group.")
                return
            self.delete_group(idx)
            reload_tree()

        def on_members() -> None:
            idx = selected_group_index()
            if idx is None:
                messagebox.showwarning(self.t("info"), "请先选择要管理成员的小组！" if self.locale == "zh_CN" else "Select a group.")
                return
            self.open_member_manager(idx)
            reload_tree()

        tk.Button(btn_frame, text="✏️ 编辑" if self.config.enable_emoji else ("编辑" if self.locale == "zh_CN" else "Edit"),
                  command=on_edit, bg=self.theme.accent_blue, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🗑️ 删除" if self.config.enable_emoji else ("删除" if self.locale == "zh_CN" else "Delete"),
                  command=on_delete, bg=self.theme.accent_red, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="👥 成员" if self.config.enable_emoji else ("成员" if self.locale == "zh_CN" else "Members"),
                  command=on_members, bg=self.theme.accent_green, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(
            btn_frame,
            text=("👥 全部成员" if self.config.enable_emoji else ("全部成员" if self.locale == "zh_CN" else "All Members")),
            command=self.open_members_manager_global,
            bg=self.theme.accent_purple,
            **btn_style,
        ).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🎨 颜色检查" if self.config.enable_emoji else ("颜色检查" if self.locale == "zh_CN" else "Colors"),
                  command=self.open_color_check, bg=self.theme.accent_neutral, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="➕ 添加" if self.config.enable_emoji else ("添加" if self.locale == "zh_CN" else "Add"),
                  command=lambda: self.add_new_group(reload_tree), bg=self.theme.accent_purple, **btn_style).pack(
            side=tk.LEFT, padx=5
        )
        tk.Button(btn_frame, text="关闭" if self.locale == "zh_CN" else "Close", command=manager_win.destroy,
                  bg=self.theme.accent_neutral, **btn_style).pack(side=tk.RIGHT, padx=5)

    def open_members_manager_global(self) -> None:
        win = tk.Toplevel(self.root)
        win.title("成员管理" if self.locale == "zh_CN" else "Members")
        win.geometry("860x520")
        win.configure(bg=self.theme.bg_primary)
        win.transient(self.root)
        win.grab_set()

        main_frame = tk.Frame(win, bg=self.theme.bg_primary, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        list_frame = tk.LabelFrame(
            main_frame,
            text="成员列表" if self.locale == "zh_CN" else "Member List",
            font=self.font_title,
            bg=self.theme.bg_primary,
            fg=self.theme.text_primary,
        )
        list_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("name", "role", "group")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=14)
        tree.heading("name", text="姓名" if self.locale == "zh_CN" else "Name")
        tree.heading("role", text="角色" if self.locale == "zh_CN" else "Role")
        tree.heading("group", text="所属组" if self.locale == "zh_CN" else "Group")
        tree.column("name", width=240)
        tree.column("role", width=120, anchor=tk.CENTER)
        tree.column("group", width=260)

        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def group_names() -> list[str]:
            return [str(g.get("name", "")).strip() for g in self.data.get("groups", []) if str(g.get("name", "")).strip()]

        def parse_selected_ref() -> Optional[tuple[int, int]]:
            sel = tree.selection()
            if not sel:
                return None
            try:
                gi_s, mi_s = str(sel[0]).split(":", 1)
                return int(gi_s), int(mi_s)
            except Exception:
                return None

        def reload_tree() -> None:
            for item in tree.get_children():
                tree.delete(item)
            for gi, g in enumerate(self.data.get("groups", [])):
                gname = str(g.get("name", "")).strip()
                members = g.get("members", [])
                if not isinstance(members, list):
                    continue
                for mi, m in enumerate(members):
                    tree.insert(
                        "",
                        tk.END,
                        iid=f"{gi}:{mi}",
                        values=(str(m.get("name", "")).strip(), str(m.get("role", "组员")).strip(), gname),
                    )

        def ensure_excel_ready() -> Any:
            try:
                import openpyxl  # type: ignore

                return openpyxl
            except Exception:
                messagebox.showerror(
                    self.t("err"),
                    "缺少 openpyxl，无法导入/导出 Excel。\n"
                    "开发环境请安装：pip install openpyxl\n"
                    "如已打包为 exe，请在打包环境安装 openpyxl 后重新打包。"
                    if self.locale == "zh_CN"
                    else "Missing openpyxl. Install: pip install openpyxl. Rebuild exe after installing.",
                    parent=win,
                )
                return None

        def export_excel() -> None:
            openpyxl = ensure_excel_ready()
            if openpyxl is None:
                return
            path = filedialog.asksaveasfilename(
                parent=win,
                title="导出 Excel" if self.locale == "zh_CN" else "Export Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
            )
            if not path:
                return

            wb = openpyxl.Workbook()
            ws_groups = wb.active
            ws_groups.title = "Groups"
            ws_groups.append(["name", "score", "color", "badge", "pattern"])
            for g in self.data.get("groups", []):
                ws_groups.append(
                    [
                        str(g.get("name", "")).strip(),
                        float(g.get("score", 0.0)),
                        str(_extract_hex(g.get("color", "")) or "").lower(),
                        str(g.get("badge", "")).strip(),
                        str(g.get("pattern", "")).strip(),
                    ]
                )

            ws_members = wb.create_sheet("Members")
            ws_members.append(["group", "name", "role"])
            for g in self.data.get("groups", []):
                gname = str(g.get("name", "")).strip()
                members = g.get("members", [])
                if not isinstance(members, list):
                    continue
                for m in members:
                    ws_members.append([gname, str(m.get("name", "")).strip(), str(m.get("role", "组员")).strip()])

            try:
                wb.save(path)
            except Exception as e:
                messagebox.showerror(self.t("err"), str(e))
                return
            messagebox.showinfo(self.t("info"), "导出完成。" if self.locale == "zh_CN" else "Exported.")

        def import_excel() -> None:
            openpyxl = ensure_excel_ready()
            if openpyxl is None:
                return
            path = filedialog.askopenfilename(
                parent=win,
                title="导入 Excel" if self.locale == "zh_CN" else "Import Excel",
                filetypes=[("Excel", "*.xlsx")],
            )
            if not path:
                return
            clear_first = messagebox.askyesno(
                self.t("info"),
                "导入前是否清空现有小组与成员？" if self.locale == "zh_CN" else "Clear existing groups/members before import?",
                parent=win,
            )
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
            except Exception as e:
                messagebox.showerror(self.t("err"), str(e))
                return

            sheet_map = {str(name).strip().lower(): name for name in wb.sheetnames}
            groups_sheet_name = sheet_map.get("groups")
            members_sheet_name = sheet_map.get("members")
            if not groups_sheet_name and not members_sheet_name:
                messagebox.showerror(
                    self.t("err"),
                    "未找到工作表：Groups / Members" if self.locale == "zh_CN" else "Missing sheets: Groups / Members",
                )
                return

            if clear_first:
                self.data["groups"] = []

            def get_or_create_group(name: str) -> dict[str, Any]:
                for g in self.data.get("groups", []):
                    if str(g.get("name", "")).strip() == name:
                        if "members" not in g or not isinstance(g.get("members"), list):
                            g["members"] = []
                        return g
                used_hex = {c.lower() for gg in self.data.get("groups", []) for c in [(_extract_hex(gg.get("color", "")) or "")] if c}
                preferred = _next_palette_entry(used_hex)
                group = {
                    "name": name,
                    "score": 0.0,
                    "color": str(preferred["hex"]).lower(),
                    "badge": str(preferred["badge"]),
                    "pattern": str(preferred["pattern"]),
                    "members": [],
                }
                self.data["groups"].append(group)
                return group

            if groups_sheet_name:
                ws = wb[groups_sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    name = str(row[0]).strip() if row[0] is not None else ""
                    if not name:
                        continue
                    g = get_or_create_group(name)
                    if len(row) > 1 and row[1] is not None and str(row[1]).strip() != "":
                        try:
                            g["score"] = float(row[1])
                        except Exception:
                            pass
                    if len(row) > 2 and row[2] is not None and str(row[2]).strip():
                        g["color"] = str(row[2]).strip().lower()
                    if len(row) > 3 and row[3] is not None and str(row[3]).strip():
                        g["badge"] = str(row[3]).strip()
                    if len(row) > 4 and row[4] is not None and str(row[4]).strip():
                        g["pattern"] = str(row[4]).strip()

            if members_sheet_name:
                ws = wb[members_sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    gname = str(row[0]).strip() if row[0] is not None else ""
                    mname = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                    role = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "组员"
                    if not gname or not mname:
                        continue
                    g = get_or_create_group(gname)
                    members = g.setdefault("members", [])
                    existing = next((m for m in members if str(m.get("name", "")).strip() == mname), None)
                    if existing is None:
                        members.append({"name": mname, "role": role or "组员"})
                    else:
                        existing["role"] = role or str(existing.get("role", "组员"))

            self._scores_dirty = True
            self._ranking_dirty = True
            self._history_dirty = True
            self._mark_dirty()
            self._rebuild_cards()
            self.refresh_all()
            reload_tree()
            messagebox.showinfo(self.t("info"), "导入完成。" if self.locale == "zh_CN" else "Imported.")

        reload_tree()

        btn_frame = tk.Frame(main_frame, bg=self.theme.bg_primary)
        btn_frame.pack(fill=tk.X, pady=12)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 14, "pady": 8, "fg": "white"}

        def on_add() -> None:
            dialog = tk.Toplevel(win)
            dialog.title("添加成员" if self.locale == "zh_CN" else "Add Member")
            dialog.geometry("520x280")
            dialog.configure(bg=self.theme.bg_primary)
            dialog.transient(win)
            dialog.grab_set()

            title_frame = tk.Frame(dialog, bg=self.theme.accent_blue, height=52)
            title_frame.pack(fill=tk.X)
            title_frame.pack_propagate(False)
            tk.Label(
                title_frame,
                text="➕ 添加新成员" if self.locale == "zh_CN" else "Add Member",
                font=self.font_header,
                bg=self.theme.accent_blue,
                fg="white",
                pady=12,
            ).pack(fill=tk.X)

            frame = tk.Frame(dialog, bg=self.theme.bg_primary, padx=24, pady=18)
            frame.pack(fill=tk.BOTH, expand=True)

            tk.Label(frame, text="姓名：" if self.locale == "zh_CN" else "Name:", font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(row=0, column=0, sticky="w", pady=10)
            name_entry = tk.Entry(frame, width=30, font=self.font_body, bg=self.theme.bg_secondary, fg=self.theme.text_primary, bd=1, relief=tk.SOLID, highlightbackground=self.theme.border, highlightthickness=1)
            name_entry.grid(row=0, column=1, sticky="ew", pady=10, padx=10)

            tk.Label(frame, text="角色：" if self.locale == "zh_CN" else "Role:", font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(row=1, column=0, sticky="w", pady=10)
            role_var = tk.StringVar(value="组员")
            role_combo = ttk.Combobox(frame, textvariable=role_var, values=["组长", "组员"], state="readonly", width=27)
            role_combo.grid(row=1, column=1, sticky="ew", pady=10, padx=10)

            tk.Label(frame, text="所属组：" if self.locale == "zh_CN" else "Group:", font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(row=2, column=0, sticky="w", pady=10)
            gvar = tk.StringVar(value=(group_names()[0] if group_names() else ""))
            gcombo = ttk.Combobox(frame, textvariable=gvar, values=group_names(), state="readonly", width=27)
            gcombo.grid(row=2, column=1, sticky="ew", pady=10, padx=10)

            def on_save() -> None:
                name = name_entry.get().strip()
                gname = gvar.get().strip()
                if not name:
                    messagebox.showerror(self.t("err"), "请输入姓名！" if self.locale == "zh_CN" else "Enter a name.", parent=dialog)
                    return
                if not gname:
                    messagebox.showerror(self.t("err"), "请选择所属组！" if self.locale == "zh_CN" else "Select a group.", parent=dialog)
                    return
                idx = None
                for i, g in enumerate(self.data.get("groups", [])):
                    if str(g.get("name", "")).strip() == gname:
                        idx = i
                        break
                if idx is None:
                    messagebox.showerror(self.t("err"), "所属组不存在。" if self.locale == "zh_CN" else "Group not found.", parent=dialog)
                    return
                group = self.data["groups"][idx]
                if "members" not in group or not isinstance(group.get("members"), list):
                    group["members"] = []
                if any(str(m.get("name", "")).strip() == name for m in group["members"]):
                    messagebox.showerror(self.t("err"), "该小组已存在同名成员！" if self.locale == "zh_CN" else "Duplicate member.", parent=dialog)
                    return
                group["members"].append({"name": name, "role": role_var.get()})
                self.add_history(gname, f"新增成员【{name}】")
                self._mark_dirty()
                self.refresh_all()
                reload_tree()
                dialog.destroy()

            bottom = tk.Frame(frame, bg=self.theme.bg_primary)
            bottom.grid(row=3, column=0, columnspan=2, pady=18)
            bstyle = {"font": self.font_body, "bd": 0, "padx": 22, "pady": 8, "fg": "white"}
            tk.Button(bottom, text="添加" if self.locale == "zh_CN" else "Add", command=on_save, bg=self.theme.accent_green, **bstyle).pack(side=tk.LEFT, padx=12)
            tk.Button(bottom, text="取消" if self.locale == "zh_CN" else "Cancel", command=dialog.destroy, bg=self.theme.accent_neutral, **bstyle).pack(side=tk.LEFT, padx=12)

            frame.columnconfigure(1, weight=1)

        def on_edit() -> None:
            ref = parse_selected_ref()
            if ref is None:
                messagebox.showwarning(self.t("info"), "请先选择成员！" if self.locale == "zh_CN" else "Select a member.", parent=win)
                return
            gi, mi = ref
            if gi >= len(self.data.get("groups", [])):
                return
            group = self.data["groups"][gi]
            members = group.get("members", [])
            if not isinstance(members, list) or mi >= len(members):
                return
            member = members[mi]
            src_group_name = str(group.get("name", "")).strip()

            dialog = tk.Toplevel(win)
            dialog.title("编辑成员" if self.locale == "zh_CN" else "Edit Member")
            dialog.geometry("520x300")
            dialog.configure(bg=self.theme.bg_primary)
            dialog.transient(win)
            dialog.grab_set()

            title_frame = tk.Frame(dialog, bg=self.theme.accent_blue, height=52)
            title_frame.pack(fill=tk.X)
            title_frame.pack_propagate(False)
            tk.Label(
                title_frame,
                text="✏️ 编辑成员信息" if self.locale == "zh_CN" else "Edit Member",
                font=self.font_header,
                bg=self.theme.accent_blue,
                fg="white",
                pady=12,
            ).pack(fill=tk.X)

            frame = tk.Frame(dialog, bg=self.theme.bg_primary, padx=24, pady=18)
            frame.pack(fill=tk.BOTH, expand=True)

            tk.Label(frame, text="姓名：" if self.locale == "zh_CN" else "Name:", font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(row=0, column=0, sticky="w", pady=10)
            name_entry = tk.Entry(frame, width=30, font=self.font_body, bg=self.theme.bg_secondary, fg=self.theme.text_primary, bd=1, relief=tk.SOLID, highlightbackground=self.theme.border, highlightthickness=1)
            name_entry.insert(0, str(member.get("name", "")).strip())
            name_entry.grid(row=0, column=1, sticky="ew", pady=10, padx=10)

            tk.Label(frame, text="角色：" if self.locale == "zh_CN" else "Role:", font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(row=1, column=0, sticky="w", pady=10)
            role_var = tk.StringVar(value=str(member.get("role", "组员")).strip() or "组员")
            role_combo = ttk.Combobox(frame, textvariable=role_var, values=["组长", "组员"], state="readonly", width=27)
            role_combo.grid(row=1, column=1, sticky="ew", pady=10, padx=10)

            tk.Label(frame, text="所属组：" if self.locale == "zh_CN" else "Group:", font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(row=2, column=0, sticky="w", pady=10)
            gvar = tk.StringVar(value=src_group_name)
            gcombo = ttk.Combobox(frame, textvariable=gvar, values=group_names(), state="readonly", width=27)
            gcombo.grid(row=2, column=1, sticky="ew", pady=10, padx=10)

            def on_save() -> None:
                new_name = name_entry.get().strip()
                new_role = role_var.get().strip() or "组员"
                dest_group_name = gvar.get().strip()
                if not new_name:
                    messagebox.showerror(self.t("err"), "请输入姓名！" if self.locale == "zh_CN" else "Enter a name.", parent=dialog)
                    return
                if not dest_group_name:
                    messagebox.showerror(self.t("err"), "请选择所属组！" if self.locale == "zh_CN" else "Select a group.", parent=dialog)
                    return
                dest_gi = None
                for i, g in enumerate(self.data.get("groups", [])):
                    if str(g.get("name", "")).strip() == dest_group_name:
                        dest_gi = i
                        break
                if dest_gi is None:
                    messagebox.showerror(self.t("err"), "所属组不存在。" if self.locale == "zh_CN" else "Group not found.", parent=dialog)
                    return
                if dest_gi == gi:
                    for i, m in enumerate(members):
                        if i != mi and str(m.get("name", "")).strip() == new_name:
                            messagebox.showerror(self.t("err"), "该小组已存在同名成员！" if self.locale == "zh_CN" else "Duplicate member.", parent=dialog)
                            return
                    old_name = str(member.get("name", "")).strip()
                    member["name"] = new_name
                    member["role"] = new_role
                    if old_name != new_name:
                        self.add_history(src_group_name, f"成员重命名：{old_name} → {new_name}")
                    else:
                        self.add_history(src_group_name, f"编辑成员【{new_name}】")
                else:
                    dest_group = self.data["groups"][dest_gi]
                    dest_members = dest_group.setdefault("members", [])
                    if any(str(m.get("name", "")).strip() == new_name for m in dest_members):
                        messagebox.showerror(self.t("err"), "目标小组已存在同名成员。" if self.locale == "zh_CN" else "Duplicate member.", parent=dialog)
                        return
                    moved = {"name": new_name, "role": new_role}
                    dest_members.append(moved)
                    del members[mi]
                    self.add_history(src_group_name, f"成员【{new_name}】转移到【{dest_group_name}】")
                    self.add_history(dest_group_name, f"接收成员【{new_name}】（来自【{src_group_name}】）")

                self._mark_dirty()
                self.refresh_all()
                reload_tree()
                dialog.destroy()

            bottom = tk.Frame(frame, bg=self.theme.bg_primary)
            bottom.grid(row=3, column=0, columnspan=2, pady=18)
            bstyle = {"font": self.font_body, "bd": 0, "padx": 22, "pady": 8, "fg": "white"}
            tk.Button(bottom, text="保存" if self.locale == "zh_CN" else "Save", command=on_save, bg=self.theme.accent_green, **bstyle).pack(side=tk.LEFT, padx=12)
            tk.Button(bottom, text="取消" if self.locale == "zh_CN" else "Cancel", command=dialog.destroy, bg=self.theme.accent_neutral, **bstyle).pack(side=tk.LEFT, padx=12)

            frame.columnconfigure(1, weight=1)

        def on_delete() -> None:
            ref = parse_selected_ref()
            if ref is None:
                messagebox.showwarning(self.t("info"), "请先选择成员！" if self.locale == "zh_CN" else "Select a member.", parent=win)
                return
            gi, mi = ref
            if gi >= len(self.data.get("groups", [])):
                return
            group = self.data["groups"][gi]
            members = group.get("members", [])
            if not isinstance(members, list) or mi >= len(members):
                return
            member_name = str(members[mi].get("name", "")).strip()
            gname = str(group.get("name", "")).strip()
            msg = f"确定要删除成员 '{member_name}' 吗？\n此操作不可恢复！" if self.locale == "zh_CN" else f"Delete '{member_name}'?"
            if not messagebox.askyesno(self.t("confirm"), msg, parent=win):
                return
            members.pop(mi)
            self.add_history(gname, f"删除成员【{member_name}】")
            self._mark_dirty()
            self.refresh_all()
            reload_tree()

        tk.Button(btn_frame, text="➕ 添加" if self.locale == "zh_CN" else "Add", command=on_add, bg=self.theme.accent_green, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✏️ 编辑" if self.locale == "zh_CN" else "Edit", command=on_edit, bg=self.theme.accent_blue, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🗑️ 删除" if self.locale == "zh_CN" else "Delete", command=on_delete, bg=self.theme.accent_red, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="⬆️ 导出Excel" if self.locale == "zh_CN" else "Export Excel", command=export_excel, bg=self.theme.accent_purple, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="⬇️ 导入Excel" if self.locale == "zh_CN" else "Import Excel", command=import_excel, bg=self.theme.accent_neutral, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="关闭" if self.locale == "zh_CN" else "Close", command=win.destroy, bg=self.theme.accent_neutral, **btn_style).pack(side=tk.RIGHT, padx=5)

    def open_color_check(self) -> None:
        win = tk.Toplevel(self.root)
        win.title("颜色检查" if self.locale == "zh_CN" else "Color Check")
        win.geometry("860x460")
        win.configure(bg=self.theme.bg_primary)
        win.transient(self.root)
        win.grab_set()

        top = tk.Frame(win, bg=self.theme.bg_primary, padx=12, pady=12)
        top.pack(fill=tk.BOTH, expand=True)

        info = tk.Label(
            top,
            text="用于快速发现：颜色重复、可读性偏弱（对比度不足）、自定义颜色占比。建议在此基础上做真实用户识别测试。"
            if self.locale == "zh_CN"
            else "Quick checks: duplicates, low contrast, custom colors. Run real user recognition tests too.",
            bg=self.theme.bg_primary,
            fg=self.theme.text_secondary,
            font=self.font_small,
            wraplength=820,
            justify="left",
        )
        info.pack(anchor="w", pady=(0, 10))

        columns = ("name", "badge", "color", "contrast", "dup", "source")
        tree = ttk.Treeview(top, columns=columns, show="headings", height=14)
        tree.heading("name", text="小组" if self.locale == "zh_CN" else "Group")
        tree.heading("badge", text="标识" if self.locale == "zh_CN" else "Badge")
        tree.heading("color", text="颜色" if self.locale == "zh_CN" else "Color")
        tree.heading("contrast", text="对比度" if self.locale == "zh_CN" else "Contrast")
        tree.heading("dup", text="重复" if self.locale == "zh_CN" else "Duplicate")
        tree.heading("source", text="来源" if self.locale == "zh_CN" else "Source")
        tree.column("name", width=240)
        tree.column("badge", width=70, anchor=tk.CENTER)
        tree.column("color", width=120, anchor=tk.CENTER)
        tree.column("contrast", width=120, anchor=tk.CENTER)
        tree.column("dup", width=80, anchor=tk.CENTER)
        tree.column("source", width=120, anchor=tk.CENTER)

        sb = ttk.Scrollbar(top, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        seen: set[str] = set()
        for g in self.data.get("groups", []):
            name = str(g.get("name", ""))
            badge = str(g.get("badge", ""))
            hex_color = _extract_hex(g.get("color", "")) or GROUP_COLOR_PALETTE[0]["hex"]
            fg = _best_text_color(hex_color)
            ratio = _contrast_ratio(fg, hex_color)
            is_dup = hex_color.lower() in seen
            dup = ("是" if is_dup else "否") if self.locale == "zh_CN" else ("Yes" if is_dup else "No")
            seen.add(hex_color.lower())
            in_palette = any(p["hex"].lower() == hex_color.lower() for p in GROUP_COLOR_PALETTE)
            source = ("推荐" if in_palette else "自定义") if self.locale == "zh_CN" else ("Palette" if in_palette else "Custom")
            tree.insert("", tk.END, values=(name, badge, hex_color.lower(), f"{ratio:.1f}", dup, source))

        bottom = tk.Frame(win, bg=self.theme.bg_primary, padx=12, pady=10)
        bottom.pack(fill=tk.X)
        tk.Button(bottom, text="关闭" if self.locale == "zh_CN" else "Close", font=self.font_body, command=win.destroy).pack(side=tk.RIGHT)

    def add_new_group(self, on_updated: Optional[Callable[[], None]]) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title(self.t("menu_add_group"))
        dialog.geometry("520x260")
        dialog.configure(bg=self.theme.bg_primary)
        dialog.transient(self.root)
        dialog.grab_set()

        title_frame = tk.Frame(dialog, bg=self.theme.accent_blue, height=52)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="➕ 添加新小组" if self.locale == "zh_CN" else "Add Group",
                 font=self.font_header, bg=self.theme.accent_blue, fg="white", pady=12).pack(fill=tk.X)

        frame = tk.Frame(dialog, bg=self.theme.bg_primary, padx=24, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="小组名称：" if self.locale == "zh_CN" else "Name:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=0, column=0, sticky="w", pady=10
        )
        name_entry = tk.Entry(frame, width=30, font=self.font_body, bg=self.theme.bg_secondary, fg=self.theme.text_primary,
                              bd=1, relief=tk.SOLID, highlightbackground=self.theme.border, highlightthickness=1)
        name_entry.grid(row=0, column=1, pady=10, padx=10, sticky="ew")

        tk.Label(frame, text="小组颜色：" if self.locale == "zh_CN" else "Color:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=1, column=0, sticky="w", pady=10
        )
        used_hex = {c.lower() for g in self.data.get("groups", []) for c in [(_extract_hex(g.get("color", "")) or "")] if c}
        preferred = _next_palette_entry(used_hex)
        color_hex_var = tk.StringVar(value=str(preferred["hex"]).lower())
        color_line = tk.Frame(frame, bg=self.theme.bg_primary)
        color_line.grid(row=1, column=1, pady=10, padx=10, sticky="ew")

        color_swatch = tk.Canvas(color_line, width=40, height=24, highlightthickness=0, bd=0)
        color_swatch.pack(side=tk.LEFT)

        def sync_swatch() -> None:
            self._draw_swatch(color_swatch, color=str(color_hex_var.get()), pattern=str(preferred["pattern"]), badge=str(preferred["badge"]))

        sync_swatch()

        color_hex_label = tk.Label(
            color_line,
            textvariable=color_hex_var,
            bg=self.theme.bg_primary,
            fg=self.theme.text_primary,
            font=self.font_body,
        )
        color_hex_label.pack(side=tk.LEFT, padx=10)

        def on_pick_color() -> None:
            current = _extract_hex(color_hex_var.get()) or preferred["hex"]
            chosen = colorchooser.askcolor(color=current, parent=dialog, title="选择颜色")
            if not chosen or not chosen[1]:
                return
            color_hex_var.set(str(chosen[1]).lower())
            sync_swatch()

        tk.Button(color_line, text="选择颜色…", font=self.font_body, bd=0, padx=10, pady=6,
                  bg=self.theme.accent_neutral, fg="white", command=on_pick_color).pack(side=tk.LEFT, padx=(8, 0))
        color_swatch.bind("<Button-1>", lambda _e: on_pick_color())
        color_hex_label.bind("<Button-1>", lambda _e: on_pick_color())

        def on_add() -> None:
            name = name_entry.get().strip()
            if not name:
                messagebox.showerror(self.t("err"), "请输入小组名称！" if self.locale == "zh_CN" else "Enter a name.")
                return
            for g in self.data.get("groups", []):
                if g.get("name") == name:
                    messagebox.showerror(self.t("err"), "该小组名称已存在！" if self.locale == "zh_CN" else "Name already exists.")
                    return
            used_hex_local = {c.lower() for g in self.data.get("groups", []) for c in [(_extract_hex(g.get("color", "")) or "")] if c}
            used_badges = {str(g.get("badge", "")).strip() for g in self.data.get("groups", []) if str(g.get("badge", "")).strip()}
            picked_hex = _extract_hex(color_hex_var.get()) or GROUP_COLOR_PALETTE[0]["hex"]
            if picked_hex.lower() in used_hex_local:
                picked_hex = _next_palette_entry(used_hex_local)["hex"]

            entry = next((p for p in GROUP_COLOR_PALETTE if p["hex"].lower() == picked_hex.lower()), None)
            if entry is None:
                entry = next((p for p in GROUP_COLOR_PALETTE if p["badge"] not in used_badges), None) or GROUP_COLOR_PALETTE[0]
            badge = entry["badge"]
            if badge in used_badges:
                badge = f"{badge}{len(used_badges) + 1}"
            pattern = entry["pattern"]

            self.data["groups"].append(
                {"name": name, "score": 0.0, "color": picked_hex.lower(), "badge": badge, "pattern": pattern, "members": []}
            )
            self._history_dirty = True
            self._scores_dirty = True
            self._ranking_dirty = True
            self._mark_dirty()
            self._rebuild_cards()
            if on_updated:
                on_updated()
            dialog.destroy()

        btn_frame = tk.Frame(frame, bg=self.theme.bg_primary)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=18)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 22, "pady": 8, "fg": "white"}
        tk.Button(btn_frame, text="添加" if self.locale == "zh_CN" else "Add", command=on_add,
                  bg=self.theme.accent_green, **btn_style).pack(side=tk.LEFT, padx=12)
        tk.Button(btn_frame, text="取消" if self.locale == "zh_CN" else "Cancel", command=dialog.destroy,
                  bg=self.theme.accent_neutral, **btn_style).pack(side=tk.LEFT, padx=12)

        frame.columnconfigure(1, weight=1)

    def edit_group(self, group_idx: int, on_updated: Optional[Callable[[], None]] = None) -> None:
        group = self.data["groups"][group_idx]
        dialog = tk.Toplevel(self.root)
        dialog.title("编辑小组" if self.locale == "zh_CN" else "Edit Group")
        dialog.geometry("520x260")
        dialog.configure(bg=self.theme.bg_primary)
        dialog.transient(self.root)
        dialog.grab_set()

        title_frame = tk.Frame(dialog, bg=self.theme.accent_blue, height=52)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="✏️ 编辑小组" if self.locale == "zh_CN" else "Edit Group",
                 font=self.font_header, bg=self.theme.accent_blue, fg="white", pady=12).pack(fill=tk.X)

        frame = tk.Frame(dialog, bg=self.theme.bg_primary, padx=24, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="小组名称：" if self.locale == "zh_CN" else "Name:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=0, column=0, sticky="w", pady=10
        )
        name_entry = tk.Entry(frame, width=30, font=self.font_body, bg=self.theme.bg_secondary, fg=self.theme.text_primary,
                              bd=1, relief=tk.SOLID, highlightbackground=self.theme.border, highlightthickness=1)
        name_entry.insert(0, str(group.get("name", "")))
        name_entry.grid(row=0, column=1, pady=10, padx=10, sticky="ew")

        tk.Label(frame, text="小组颜色：" if self.locale == "zh_CN" else "Color:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=1, column=0, sticky="w", pady=10
        )
        current_hex = _extract_hex(group.get("color", "")) or GROUP_COLOR_PALETTE[0]["hex"]
        color_hex_var = tk.StringVar(value=str(current_hex).lower())
        color_line = tk.Frame(frame, bg=self.theme.bg_primary)
        color_line.grid(row=1, column=1, pady=10, padx=10, sticky="ew")

        color_swatch = tk.Canvas(color_line, width=40, height=24, highlightthickness=0, bd=0)
        color_swatch.pack(side=tk.LEFT)

        def sync_swatch() -> None:
            badge = str(group.get("badge", "")).strip() or "?"
            pattern = str(group.get("pattern", "diag")).strip() or "diag"
            self._draw_swatch(color_swatch, color=str(color_hex_var.get()), pattern=pattern, badge=badge)

        sync_swatch()

        color_hex_label = tk.Label(
            color_line,
            textvariable=color_hex_var,
            bg=self.theme.bg_primary,
            fg=self.theme.text_primary,
            font=self.font_body,
        )
        color_hex_label.pack(side=tk.LEFT, padx=10)

        def on_pick_color() -> None:
            current = _extract_hex(color_hex_var.get()) or current_hex
            chosen = colorchooser.askcolor(color=current, parent=dialog, title="选择颜色")
            if not chosen or not chosen[1]:
                return
            color_hex_var.set(str(chosen[1]).lower())
            sync_swatch()

        tk.Button(color_line, text="选择颜色…", font=self.font_body, bd=0, padx=10, pady=6,
                  bg=self.theme.accent_neutral, fg="white", command=on_pick_color).pack(side=tk.LEFT, padx=(8, 0))
        color_swatch.bind("<Button-1>", lambda _e: on_pick_color())
        color_hex_label.bind("<Button-1>", lambda _e: on_pick_color())

        def on_save() -> None:
            new_name = name_entry.get().strip()
            if not new_name:
                messagebox.showerror(self.t("err"), "请输入小组名称！" if self.locale == "zh_CN" else "Enter a name.")
                return
            for i, g in enumerate(self.data.get("groups", [])):
                if i != group_idx and g.get("name") == new_name:
                    messagebox.showerror(self.t("err"), "该小组名称已存在！" if self.locale == "zh_CN" else "Name already exists.")
                    return
            used_hex_local = {
                (_extract_hex(g.get("color", "")) or "").lower()
                for i, g in enumerate(self.data.get("groups", []))
                if i != group_idx and (_extract_hex(g.get("color", "")) or "")
            }
            picked_hex = _extract_hex(color_hex_var.get()) or GROUP_COLOR_PALETTE[0]["hex"]
            if picked_hex.lower() in used_hex_local:
                picked_hex = _next_palette_entry(used_hex_local)["hex"]
                messagebox.showinfo(
                    self.t("info"),
                    f"所选颜色已被其他小组使用，已自动调整为：{picked_hex}" if self.locale == "zh_CN" else f"Color already used. Switched to {picked_hex}",
                )
            old_hex = _extract_hex(self.data["groups"][group_idx].get("color", "")) or current_hex
            self.data["groups"][group_idx]["name"] = new_name
            self.data["groups"][group_idx]["color"] = picked_hex.lower()
            if "badge" not in self.data["groups"][group_idx] or not str(self.data["groups"][group_idx].get("badge", "")).strip():
                used_badges = {
                    str(g.get("badge", "")).strip()
                    for i, g in enumerate(self.data.get("groups", []))
                    if i != group_idx and str(g.get("badge", "")).strip()
                }
                entry = next((p for p in GROUP_COLOR_PALETTE if p["hex"].lower() == picked_hex.lower()), None) or GROUP_COLOR_PALETTE[0]
                badge = entry["badge"]
                if badge in used_badges:
                    badge = f"{badge}{len(used_badges) + 1}"
                self.data["groups"][group_idx]["badge"] = badge
            if "pattern" not in self.data["groups"][group_idx] or not str(self.data["groups"][group_idx].get("pattern", "")).strip():
                entry = next((p for p in GROUP_COLOR_PALETTE if p["hex"].lower() == picked_hex.lower()), None) or GROUP_COLOR_PALETTE[0]
                self.data["groups"][group_idx]["pattern"] = entry["pattern"]
            self._ranking_dirty = True
            self._mark_dirty()
            if self.batch_group_combo and self.batch_group_var:
                values = [self.t("all_groups")] + [g.get("name", "") for g in self.data.get("groups", [])]
                self.batch_group_combo.configure(values=values)
                if self.batch_group_var.get() not in values:
                    self.batch_group_var.set(self.t("all_groups"))

            if getattr(self, "group_cards", None) and self._cards_ready:
                card_info = next((ci for ci in self.group_cards if int(ci.get("group_idx", -1)) == group_idx), None)
                if card_info is not None:
                    new_hex = str(self.data["groups"][group_idx].get("color", "")).lower()
                    if str(old_hex).lower() != new_hex:
                        self._animate_group_card_color(card_info, from_color=str(old_hex).lower(), to_color=new_hex)
                    else:
                        self._apply_group_color_to_card(card_info, color=new_hex)
                else:
                    self._rebuild_cards()
            else:
                self._rebuild_cards()
            self.refresh_ranking()
            if on_updated:
                on_updated()
            dialog.destroy()

        btn_frame = tk.Frame(frame, bg=self.theme.bg_primary)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=18)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 22, "pady": 8, "fg": "white"}
        tk.Button(btn_frame, text="保存" if self.locale == "zh_CN" else "Save", command=on_save,
                  bg=self.theme.accent_blue, **btn_style).pack(side=tk.LEFT, padx=12)
        tk.Button(btn_frame, text="取消" if self.locale == "zh_CN" else "Cancel", command=dialog.destroy,
                  bg=self.theme.accent_neutral, **btn_style).pack(side=tk.LEFT, padx=12)

        frame.columnconfigure(1, weight=1)

    def delete_group(self, group_idx: int) -> None:
        group_name = str(self.data["groups"][group_idx].get("name", ""))
        msg = f"确定要删除小组 '{group_name}' 吗？\n此操作不可恢复！" if self.locale == "zh_CN" else f"Delete '{group_name}'?"
        if messagebox.askyesno(self.t("confirm"), msg):
            self.data["groups"].pop(group_idx)
            self.add_history("系统" if self.locale == "zh_CN" else "System",
                             f"删除小组：{group_name}" if self.locale == "zh_CN" else f"Deleted group: {group_name}")
            self._scores_dirty = True
            self._ranking_dirty = True
            self._mark_dirty()
            self._rebuild_cards()
            messagebox.showinfo(self.t("success"), f"小组 '{group_name}' 已删除！" if self.locale == "zh_CN" else "Deleted.")

    def open_member_manager(self, group_idx: int) -> None:
        group = self.data["groups"][group_idx]
        if "members" not in group or not isinstance(group.get("members"), list):
            group["members"] = []

        dialog = tk.Toplevel(self.root)
        dialog.title(("管理成员 - " if self.locale == "zh_CN" else "Members - ") + str(group.get("name", "")))
        dialog.geometry("760x520")
        dialog.configure(bg=self.theme.bg_primary)
        dialog.transient(self.root)
        dialog.grab_set()

        title_bg = str(group.get("color", self.theme.accent_blue))
        title_frame = tk.Frame(dialog, bg=title_bg, height=58)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        title_text = f"👥 管理小组成员 - {group.get('name','')}" if self.locale == "zh_CN" else f"Members - {group.get('name','')}"
        tk.Label(title_frame, text=title_text, font=self.font_header, bg=title_bg, fg=_best_text_color(title_bg), pady=14).pack(fill=tk.X)

        main_frame = tk.Frame(dialog, bg=self.theme.bg_primary, padx=16, pady=16)
        main_frame.pack(fill=tk.BOTH, expand=True)

        list_frame = tk.LabelFrame(main_frame, text="成员列表" if self.locale == "zh_CN" else "Member List",
                                   font=self.font_title, bg=self.theme.bg_primary, fg=self.theme.text_primary)
        list_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("name", "role")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=12)
        tree.heading("name", text="姓名" if self.locale == "zh_CN" else "Name")
        tree.heading("role", text="角色" if self.locale == "zh_CN" else "Role")
        tree.column("name", width=360)
        tree.column("role", width=160, anchor=tk.CENTER)

        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def reload_members() -> None:
            for item in tree.get_children():
                tree.delete(item)
            for member in group.get("members", []):
                tree.insert("", tk.END, values=(member.get("name", ""), member.get("role", "组员")))

        def selected_member_index() -> Optional[int]:
            sel = tree.selection()
            if not sel:
                return None
            return tree.index(sel[0])

        reload_members()

        btn_frame = tk.Frame(main_frame, bg=self.theme.bg_primary)
        btn_frame.pack(fill=tk.X, pady=12)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 14, "pady": 8, "fg": "white"}

        def on_add() -> None:
            self.add_member(group_idx, on_updated=reload_members)

        def on_edit() -> None:
            idx = selected_member_index()
            if idx is None:
                messagebox.showwarning(self.t("info"), "请先选择要编辑的成员！" if self.locale == "zh_CN" else "Select a member.")
                return
            self.edit_member(group_idx, idx, on_updated=reload_members)

        def on_move() -> None:
            idx = selected_member_index()
            if idx is None:
                messagebox.showwarning(self.t("info"), "请先选择要更换小组的成员！" if self.locale == "zh_CN" else "Select a member.")
                return
            groups = list(self.data.get("groups", []))
            if len(groups) < 2:
                messagebox.showwarning(self.t("info"), "当前只有一个小组，无法更换。" if self.locale == "zh_CN" else "Only one group exists.")
                return
            member = group.get("members", [])[idx]
            member_name = str(member.get("name", "")).strip()
            if not member_name:
                messagebox.showerror(self.t("err"), "成员姓名为空，无法更换。" if self.locale == "zh_CN" else "Invalid member.")
                return

            dlg = tk.Toplevel(dialog)
            dlg.title("更换小组" if self.locale == "zh_CN" else "Move Member")
            dlg.geometry("440x220")
            dlg.configure(bg=self.theme.bg_primary)
            dlg.transient(dialog)
            dlg.grab_set()

            frame = tk.Frame(dlg, bg=self.theme.bg_primary, padx=18, pady=16)
            frame.pack(fill=tk.BOTH, expand=True)

            tk.Label(
                frame,
                text=f"成员：{member_name}",
                font=self.font_body,
                bg=self.theme.bg_primary,
                fg=self.theme.text_primary,
            ).pack(anchor="w", pady=(0, 12))

            tk.Label(
                frame,
                text="目标小组：",
                font=self.font_body,
                bg=self.theme.bg_primary,
                fg=self.theme.text_primary,
            ).pack(anchor="w")

            options = [str(g.get("name", "")) for i, g in enumerate(groups) if i != group_idx and str(g.get("name", "")).strip()]
            if not options:
                messagebox.showwarning(self.t("info"), "没有可选择的目标小组。" if self.locale == "zh_CN" else "No target group.")
                dlg.destroy()
                return
            target_var = tk.StringVar(value=options[0])
            cb = ttk.Combobox(frame, textvariable=target_var, values=options, state="readonly", width=28)
            cb.pack(anchor="w", pady=8)

            def do_move() -> None:
                target_name = target_var.get()
                to_idx = None
                for i, g in enumerate(groups):
                    if i != group_idx and str(g.get("name", "")) == target_name:
                        to_idx = i
                        break
                if to_idx is None:
                    messagebox.showerror(self.t("err"), "目标小组不存在。" if self.locale == "zh_CN" else "Target not found.")
                    return
                dest_group = self.data["groups"][to_idx]
                dest_members = dest_group.setdefault("members", [])
                if any(str(m.get("name", "")).strip() == member_name for m in dest_members):
                    messagebox.showerror(self.t("err"), "目标小组已存在同名成员。" if self.locale == "zh_CN" else "Duplicate member.")
                    return

                dest_members.append(member)
                del group["members"][idx]

                src_name = str(group.get("name", ""))
                self.add_history(src_name, f"成员【{member_name}】转移到【{target_name}】")
                self.add_history(target_name, f"接收成员【{member_name}】（来自【{src_name}】）")
                self._ranking_dirty = True
                self._mark_dirty()
                self.refresh_ranking()
                reload_members()
                dlg.destroy()

            btns = tk.Frame(frame, bg=self.theme.bg_primary)
            btns.pack(fill=tk.X, pady=(14, 0))
            tk.Button(btns, text="取消" if self.locale == "zh_CN" else "Cancel", command=dlg.destroy, font=self.font_body).pack(side=tk.RIGHT)
            tk.Button(btns, text="确定" if self.locale == "zh_CN" else "OK", command=do_move, font=self.font_body).pack(side=tk.RIGHT, padx=8)

        def on_delete() -> None:
            idx = selected_member_index()
            if idx is None:
                messagebox.showwarning(self.t("info"), "请先选择要删除的成员！" if self.locale == "zh_CN" else "Select a member.")
                return
            self.delete_member(group_idx, idx)
            reload_members()

        tk.Button(btn_frame, text="➕ 添加" if self.config.enable_emoji else ("添加" if self.locale == "zh_CN" else "Add"),
                  command=on_add, bg=self.theme.accent_green, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✏️ 编辑" if self.config.enable_emoji else ("编辑" if self.locale == "zh_CN" else "Edit"),
                  command=on_edit, bg=self.theme.accent_blue, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🔁 更换小组" if self.config.enable_emoji else ("更换小组" if self.locale == "zh_CN" else "Move"),
                  command=on_move, bg=self.theme.accent_neutral, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🗑️ 删除" if self.config.enable_emoji else ("删除" if self.locale == "zh_CN" else "Delete"),
                  command=on_delete, bg=self.theme.accent_red, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="关闭" if self.locale == "zh_CN" else "Close", command=dialog.destroy,
                  bg=self.theme.accent_neutral, **btn_style).pack(side=tk.RIGHT, padx=5)

    def add_member(self, group_idx: int, on_updated: Optional[Callable[[], None]] = None) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title("添加成员" if self.locale == "zh_CN" else "Add Member")
        dialog.geometry("480x280")
        dialog.configure(bg=self.theme.bg_primary)
        dialog.transient(self.root)
        dialog.grab_set()

        title_frame = tk.Frame(dialog, bg=self.theme.accent_green, height=52)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="➕ 添加新成员" if self.locale == "zh_CN" else "Add Member",
                 font=self.font_header, bg=self.theme.accent_green, fg="white", pady=12).pack(fill=tk.X)

        frame = tk.Frame(dialog, bg=self.theme.bg_primary, padx=24, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="成员姓名：" if self.locale == "zh_CN" else "Name:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=0, column=0, sticky="w", pady=10
        )
        name_entry = tk.Entry(frame, width=30, font=self.font_body, bg=self.theme.bg_secondary, fg=self.theme.text_primary,
                              bd=1, relief=tk.SOLID, highlightbackground=self.theme.border, highlightthickness=1)
        name_entry.grid(row=0, column=1, pady=10, padx=10, sticky="ew")

        tk.Label(frame, text="角色：" if self.locale == "zh_CN" else "Role:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=1, column=0, sticky="w", pady=10
        )
        role_var = tk.StringVar(value="组员")
        role_combo = ttk.Combobox(frame, textvariable=role_var, values=["组长", "副组长", "组员"], state="readonly")
        role_combo.grid(row=1, column=1, pady=10, padx=10, sticky="ew")

        def on_add() -> None:
            name = name_entry.get().strip()
            if not name:
                messagebox.showerror(self.t("err"), "请输入成员姓名！" if self.locale == "zh_CN" else "Enter a name.")
                return
            group = self.data["groups"][group_idx]
            if "members" not in group or not isinstance(group.get("members"), list):
                group["members"] = []
            for member in group.get("members", []):
                if member.get("name") == name:
                    messagebox.showerror(self.t("err"), "该成员已存在！" if self.locale == "zh_CN" else "Member exists.")
                    return
            group["members"].append({"name": name, "role": role_var.get()})
            self._mark_dirty()
            if on_updated:
                on_updated()
            dialog.destroy()

        btn_frame = tk.Frame(frame, bg=self.theme.bg_primary)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=18)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 22, "pady": 8, "fg": "white"}
        tk.Button(btn_frame, text="添加" if self.locale == "zh_CN" else "Add", command=on_add,
                  bg=self.theme.accent_green, **btn_style).pack(side=tk.LEFT, padx=12)
        tk.Button(btn_frame, text="取消" if self.locale == "zh_CN" else "Cancel", command=dialog.destroy,
                  bg=self.theme.accent_neutral, **btn_style).pack(side=tk.LEFT, padx=12)

        frame.columnconfigure(1, weight=1)

    def edit_member(self, group_idx: int, member_idx: int, on_updated: Optional[Callable[[], None]] = None) -> None:
        group = self.data["groups"][group_idx]
        if "members" not in group or not isinstance(group.get("members"), list):
            messagebox.showerror(self.t("err"), "该小组暂无成员！" if self.locale == "zh_CN" else "No members.")
            return
        member = group["members"][member_idx]

        dialog = tk.Toplevel(self.root)
        dialog.title("编辑成员" if self.locale == "zh_CN" else "Edit Member")
        dialog.geometry("480x280")
        dialog.configure(bg=self.theme.bg_primary)
        dialog.transient(self.root)
        dialog.grab_set()

        title_frame = tk.Frame(dialog, bg=self.theme.accent_blue, height=52)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="✏️ 编辑成员信息" if self.locale == "zh_CN" else "Edit Member",
                 font=self.font_header, bg=self.theme.accent_blue, fg="white", pady=12).pack(fill=tk.X)

        frame = tk.Frame(dialog, bg=self.theme.bg_primary, padx=24, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="成员姓名：" if self.locale == "zh_CN" else "Name:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=0, column=0, sticky="w", pady=10
        )
        name_entry = tk.Entry(frame, width=30, font=self.font_body, bg=self.theme.bg_secondary, fg=self.theme.text_primary,
                              bd=1, relief=tk.SOLID, highlightbackground=self.theme.border, highlightthickness=1)
        name_entry.insert(0, str(member.get("name", "")))
        name_entry.grid(row=0, column=1, pady=10, padx=10, sticky="ew")

        tk.Label(frame, text="角色：" if self.locale == "zh_CN" else "Role:",
                 font=self.font_body, bg=self.theme.bg_primary, fg=self.theme.text_primary).grid(
            row=1, column=0, sticky="w", pady=10
        )
        role_var = tk.StringVar(value=str(member.get("role", "组员")))
        role_combo = ttk.Combobox(frame, textvariable=role_var, values=["组长", "副组长", "组员"], state="readonly")
        role_combo.grid(row=1, column=1, pady=10, padx=10, sticky="ew")

        def on_save() -> None:
            new_name = name_entry.get().strip()
            if not new_name:
                messagebox.showerror(self.t("err"), "请输入成员姓名！" if self.locale == "zh_CN" else "Enter a name.")
                return
            for i, m in enumerate(group.get("members", [])):
                if i != member_idx and m.get("name") == new_name:
                    messagebox.showerror(self.t("err"), "该成员姓名已存在！" if self.locale == "zh_CN" else "Name exists.")
                    return
            group["members"][member_idx]["name"] = new_name
            group["members"][member_idx]["role"] = role_var.get()
            self._mark_dirty()
            if on_updated:
                on_updated()
            dialog.destroy()

        btn_frame = tk.Frame(frame, bg=self.theme.bg_primary)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=18)
        btn_style = {"font": self.font_body, "bd": 0, "padx": 22, "pady": 8, "fg": "white"}
        tk.Button(btn_frame, text="保存" if self.locale == "zh_CN" else "Save", command=on_save,
                  bg=self.theme.accent_blue, **btn_style).pack(side=tk.LEFT, padx=12)
        tk.Button(btn_frame, text="取消" if self.locale == "zh_CN" else "Cancel", command=dialog.destroy,
                  bg=self.theme.accent_neutral, **btn_style).pack(side=tk.LEFT, padx=12)

        frame.columnconfigure(1, weight=1)

    def delete_member(self, group_idx: int, member_idx: int) -> None:
        group = self.data["groups"][group_idx]
        if "members" not in group or not isinstance(group.get("members"), list):
            messagebox.showerror(self.t("err"), "该小组暂无成员！" if self.locale == "zh_CN" else "No members.")
            return
        member_name = str(group["members"][member_idx].get("name", ""))
        msg = f"确定要删除成员 '{member_name}' 吗？\n此操作不可恢复！" if self.locale == "zh_CN" else f"Delete '{member_name}'?"
        if messagebox.askyesno(self.t("confirm"), msg):
            group["members"].pop(member_idx)
            self._mark_dirty()
            messagebox.showinfo(self.t("success"), "已删除" if self.locale == "zh_CN" else "Deleted")

    def refresh_group_cards(self) -> None:
        if not getattr(self, "group_cards", None) or not self._cards_ready or not self._scores_dirty:
            return
        for card_info in self.group_cards:
            idx = int(card_info["group_idx"])
            if idx >= len(self.data.get("groups", [])):
                continue
            group = self.data["groups"][idx]
            score = float(group.get("score", 0.0))
            color = str(group.get("color", self.theme.accent_blue))
            card_info["score_var"].set(f"{score:.1f}")
            if not bool(card_info.get("animating")):
                self._apply_group_color_to_card(card_info, color=color)
        self._scores_dirty = False

    def refresh_ranking(self) -> None:
        if not self.rank_listbox or not self._ranking_dirty:
            return
        self.rank_listbox.delete(0, tk.END)
        sorted_groups = sorted(self.data.get("groups", []), key=lambda x: -float(x.get("score", 0.0)))
        medals = ["🥇", "🥈", "🥉"] if self.config.enable_emoji else ["1", "2", "3"]
        for i, group in enumerate(sorted_groups):
            member_count = len(group.get("members", []))
            name = group.get("name", "")
            score = float(group.get("score", 0.0))
            prefix = medals[i] if i < 3 else f"{i + 1}."
            suffix = f"({member_count}人)" if self.locale == "zh_CN" else f"({member_count})"
            tail = f"{score:.1f}分" if self.locale == "zh_CN" else f"{score:.1f}"
            self.rank_listbox.insert(tk.END, f"{prefix} {name} {suffix}：{tail}")
        self._ranking_dirty = False

    def refresh_history(self) -> None:
        if not self.history_text or not self._history_dirty:
            return
        records = list(reversed(self.data.get("history", [])))
        if self.config.history_render_limit > 0:
            records = records[: int(self.config.history_render_limit)]

        self.history_text.config(state=tk.NORMAL)
        self.history_text.delete(1.0, tk.END)
        for record in records:
            self.history_text.insert(tk.END, f"⏰ {record.get('time', '')}\n", "time")
            group = str(record.get("group", ""))
            change = record.get("change", 0)
            if isinstance(change, (int, float)):
                val = float(change)
                change_str = f"{val:+.2f}"
                tag = "pos" if val > 0 else "neg" if val < 0 else "zero"
            else:
                change_str = str(change)
                tag = "text"
            self.history_text.insert(tk.END, f"{group} {change_str}\n\n", tag)
        self.history_text.config(state=tk.DISABLED)
        self._history_dirty = False

    def refresh_all(self) -> None:
        self.refresh_group_cards()
        self.refresh_ranking()
        self.refresh_history()

    def set_theme(self, theme_name: str) -> None:
        self.config.theme = theme_name
        self.theme = resolve_theme(theme_name)
        self.root.configure(bg=self.theme.bg_primary)
        self.main_container.configure(bg=self.theme.bg_primary)
        self.left_container.configure(bg=self.theme.bg_primary)
        self.right_container.configure(bg=self.theme.bg_primary)
        self._configure_styles()
        if self.rank_listbox:
            self.rank_listbox.configure(bg=self.theme.bg_secondary, fg=self.theme.text_primary,
                                       highlightbackground=self.theme.border)
        if self.history_text:
            self.history_text.configure(bg=self.theme.bg_secondary, fg=self.theme.text_primary)
            self._configure_history_tags()
        if getattr(self, "cards_frame", None) is not None:
            self.cards_frame.configure(bg=self.theme.bg_primary)
            self._rebuild_cards()
        self._history_dirty = True
        self._ranking_dirty = True
        self._scores_dirty = True
        self.refresh_all()

    def set_font_scale(self, scale: float) -> None:
        self.config.font_scale = max(0.5, float(scale))
        self._configure_fonts()
        self._configure_styles()
        self._history_dirty = True
        self._ranking_dirty = True
        self._scores_dirty = True
        if getattr(self, "cards_frame", None) is not None:
            self._rebuild_cards()
        self.refresh_all()


def widget_tree_metrics(root: tk.Misc) -> dict[str, int]:
    def max_depth(w: tk.Misc, depth: int = 0) -> int:
        kids = w.winfo_children()
        if not kids:
            return depth
        return max(max_depth(k, depth + 1) for k in kids)

    def count_nodes(w: tk.Misc) -> int:
        return 1 + sum(count_nodes(k) for k in w.winfo_children())

    return {"max_depth": max_depth(root), "node_count": count_nodes(root)}


def run_headless_profile(tag: str, config: AppConfig) -> None:
    out_dir = os.path.join("perf_artifacts", tag)
    os.makedirs(out_dir, exist_ok=True)

    profiler = None
    try:
        from pyinstrument import Profiler

        profiler = Profiler()
        profiler.start()
    except Exception:
        profiler = None

    tracemalloc.start()
    t0 = time.perf_counter()
    root = tk.Tk()
    root.withdraw()
    t1 = time.perf_counter()
    app = GroupScoreApp(root, config=config)

    t_shell_start = time.perf_counter()
    root.update_idletasks()
    t_shell_end = time.perf_counter()

    for _ in range(400):
        root.update()
        if getattr(app, "_cards_ready", False) and getattr(app, "_right_ready", False):
            break
    t_ready = time.perf_counter()

    current, peak = tracemalloc.get_traced_memory()
    metrics = {
        "tk_init_ms": (t1 - t0) * 1000.0,
        "first_frame_ms": (t_shell_end - t_shell_start) * 1000.0,
        "ui_ready_ms": (t_ready - t1) * 1000.0,
        "tracemalloc_peak_bytes": peak,
        "layout_mode": app.layout_mode,
    }

    with open(os.path.join(out_dir, "metrics.json"), "w", encoding="utf-8") as f:
        json.dump(metrics, f, ensure_ascii=False, indent=2)

    with open(os.path.join(out_dir, "layout_tree.json"), "w", encoding="utf-8") as f:
        json.dump(widget_tree_metrics(root), f, ensure_ascii=False, indent=2)

    if profiler is not None:
        profiler.stop()
        try:
            with open(os.path.join(out_dir, "flame.html"), "w", encoding="utf-8") as f:
                f.write(profiler.output_html())
        except Exception:
            pass

    root.destroy()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--theme", choices=["light", "dark"], default="light")
    parser.add_argument("--font-scale", type=float, default=1.0)
    parser.add_argument("--locale", type=str, default=None)
    parser.add_argument("--disable-emoji", action="store_true")
    parser.add_argument("--profile", type=str, default=None)
    args = parser.parse_args()

    config = AppConfig(
        theme=args.theme,
        font_scale=args.font_scale,
        locale=args.locale,
        enable_emoji=not args.disable_emoji,
    )
    config.data_file = os.path.join(get_app_dir(), os.path.basename(config.data_file))

    if args.profile:
        run_headless_profile(args.profile, config)
        return

    root = tk.Tk()
    _ = GroupScoreApp(root, config=config)
    root.mainloop()


if __name__ == "__main__":
    main()
